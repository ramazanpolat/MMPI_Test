# -*- coding:utf-8 -*-

# 版权所有 (C) 2018.6.25 金盛羽。保留所有权利。
# Copyright 2018.6.25 Shengyu Jin. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
This module is mainly for processing functions about MMPI, including
a complete questionnaire.
"""

import time
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import Font, Alignment

# for debug
from random import randint

# Instruction
Ins1 = """-----------------------------------------------------------------                   
       Minnesota Multiphasic Per-sonality Inventory (MMPI)       
-----------------------------------------------------------------
Brief introduction: The Minnesota Multiple Personality Inventory (MMPI) 
was developed by Professor Hathaway of the University of Minnesota 37 (S.R. Hathaway) 
and McKinley (J.C.Mckinley) formulated in 1942, it can be used to test the 
personality types of normal people and can also be used to distinguish 
normal people from those with mental or mental illness. 
last world 39 At the end of the century, on the basis of years of research and 
practical investigations by Mr. Ji Shumao and others, the 40 MMPI test questionnaires and norms, 
my country began to try MMPI in 1980, 
and it has been used more widely in recent years. 
41 The results prove that this test has a certain 
reliability and validity in our country, and has a 
high clinical reference value.

-----------------------------------------------------------------
Important reminder: 
This test is a professional psychological measurement method with 
clinical reference significance, and it must be tested by relevant 
professionals. 51 Use and interpret under guidance, please use with caution!                         
-----------------------------------------------------------------"""

Ins2 = """-----------------------------------------------------------------
Instructions: This quiz consists of many questions that are relevant to you. 
As you read each question, please consider if it fits your current behaviour, 
feelings, attitude or opinion. If so, please enter "1", otherwise enter enter "0" 
and press the "Enter" key after confirmation to complete the answer to this question. 
Please fill in as soon as possible after you read the title.
Don't spend too much time thinking about each question. 
Personality is different, the answer is nothing 61 Whether it is right or wrong,
good or bad, there is no need to worry about it at all. 

Please answer according to your actual situation.
-----------------------------------------------------------------
sample answer：
x. my gender is
1. male    0. female
> 1
-----------------------------------------------------------------"""

# 566 questions of MMPI questionnaire
Que = {
 1: Teknik yazılardan hoşlanırım
2: İştahım iyidir
3: Çok defa sabahları dinç ve dinlenmiş olarak uyanırım
4: Kütüphaneci olarak çalışmayı seveceğimi sanıyorum
5: Gürültüden kolayca uyanırım
6: Cinayet haberlerini okumaktan hoşlanırım
7: Çoğu zaman el ve ayaklarımın sıcaklığı iyidir
8: Günlük hayatım beni ilgilendirecek şeylerle doludur
9: Bugün de hemen hemen eskisi kadar iyi çalışabiliyorum
10: Çoğu zaman boğazım tıkanır gibi olur
11: İnsan rüyalarını anlamaya çalışmalı ve kendini onlara göre ayarlamalıdır
12: Polis romanlarından ya da esrarengiz romanlardan hoşlanırım
13: Büyük bir sinir gerginliği içinde çalışırım
14: Ayda bir iki defa ishal olurum
15: Ara sıra söylenemeyecek kadar ayıp şeyler düşünürüm
16: Hayatta kötülükler hep beni bulur
17: Babam iyi bir adamdır
18: Pek seyrek kabız olurum
19: Yeni bir işe girince kimin gözüne girme gerektiğini öğrenmek isterim
20: Cinsel yaşamımdan memnunum
21: Zaman zaman evi bırakıp gitmek istemişimdir
22: Ara sıra kontrol edemediğim gülme ve ağlama nöbetlerine tutulurum
23: Tekrarlanan mide bulantısı ve kusmalar bana sıkıntı verir
24: Kimse beni anlamıyor
25: Şarkici olmayı isterim
26: Başım derde girince susmayı tercih ederim
27: Bazen kötü ruhların beni etkileri altına aldıklarını hissederim
28: Kötülüğe kötülükle karşılık vermem prensibimdir
29: Çoğu kez midem ekşir
30: Bazen canim küfretmek ister
31: Sık sık geceleri kabus geçiririm
32: Zihnimi bir iş üzerinde toplamada güçlük çekerim
33: Başımdan çok garip ve tuhaf şeyler geçti
34: Çoğu zaman öksürüğüm vardır
35: Başkaları engel olmasaydı daha çok başarılı olurdum
36: Sağlığım beni pek kaygılandırmaz
37: Cinsel yaşamım yüzünden başım hiç derde girmedi
38: Gençliğimde bir devre ufak tefek şeyler çaldım
39: Bazen içimde bir şeyler kırmak isteği geçer
40: Başka bir şey yapmaktansa çoğu zaman oturup hayal kurmayı severim
41: Kendimi toparlayamadığım için günler, haftalar, hatta aylarca hiçbir şeye el sürmediğim olur
42: Ailem seçtiğim veya seçmek istediğim mesleği beğenmiyor
43: Kuşkulu ve rahatsız uyurum
44: Çoğu zaman başımın her tarafı ağrır
45: Her zaman doğruyu söylemem
46: Şimdi her zamankinden daha iyi düşünüp tartabiliyorum
47: Ortada hiç bir neden yokken haftada bir ya da daha sık birdenbire her yanımı ateş basar
48: Başkaları ile bir arada iken kulağıma çok garip şeyler gelmesinden rahatsız olurum
49: Kanunların hemen hepsi kaldırılsa daha iyi olur
50: Bazen ruhum vücudumdan ayrılır
51: Sağlığım bir çok arkadaşımınki kadar iyidir
52: Uzun zamandan beri görmediğim okul arkadaşlarım ya da tanıdıklarım önce benimle konuşmazlarsa onları görmemezlikten gelmeyi tercih ederim
53: Hocaların dua okuyup üflemesi hastalığı iyileştirir
54: Tanıdıklarımın çoğu beni sever
55: Kalp ve göğüs ağrılarından hemen hemen hiç şikayetim yoktur
56: Çocukken okuldan kaçtığım için bir iki defa cezalandırıldım
57: İnsanlarla çabucak kaynaşırım
58: Kuran' ın buyurdukları bir bir çıkmaktadır
59: Çok defa benden az bilenlerden emir alarak çalışmak zorunda kaldım
60: Her gün gazetelerin baş yazılarını okumam
61: Gerektiği gibi bir hayat yaşayamadım
62: Vücudumun bazı yerlerinde çok defa yanma, gıdıklanma, karıncalanma veya uyuşukluk hissederim
63: Büyük abdest yapmada ya da tutmada hiç bir güçlük çekmem
64: Bazen başkalarının sabrını tüketecek kadar bir şeye saplanır kalırım
65: Babamı severim
66: Etrafımda başkalarının görmediği eşya, hayvanlar veya insanlar görürüm
67: Başkalarının mutlu göründüğü kadar mutlu olmayı isterdim
68: Ensemde nadiren ağrı hissederim
69: Kendi cinsimden olanları oldukça çekici bulurum
70: Körebe oyunundan hoşlanırdım
71: Birçok kimseler başkalarının ilgi ve yardımlarını sağlamak için talihsizliklerini abartırlar
72: Hemen hemen her gün mide ağrılarından rahatsız olurum
73: Ben önemli bir kimseyim
74: Çoğu zaman kız olmayı isterdim (Şayet kız iseniz:) Kız olduğuma hiç üzülmedim
75: Ara sıra öfkelenirim
76: Çoğu zaman kendimi hüzünlü hissederim
77: Aşk romanları okumaktan hoşlanırım
78: Şiiri severim
79: Kolay incinmem
80: Bazen hayvanlara rahat vermem
81: Orman bekçiliği gibi işlerden hoşlanacağımı sanıyorum
82: Tartışmalarda çabucak yenilirim
83: Çok çalışabilen ya da çalışmak isteyen kişinin başarılı olma şansı yüksektir
84: Bugünlerde artık hiç ilerleme umudum kalmamış gibi hissediyorum
85: Kullanmayacak bile olsam bazen başkalarının ayakkabı, eldiven gibi özel eşyaları o kadar hoşuma gider ki dokunmak ve aşırmak isterim
86: Kendime hiç güvenim yoktur
87: Çiçek satıcısı olmayı isterdim
88: Genel olarak hayatın yaşanmaya değer olduğu kanısındayım
89: İnsanlara gerçeği kabul ettirmek güçtür
90: Bugün yapmam gereken işleri ara sıra yarına bıraktığım olur
91: Benimle alay edilmesine aldırmam
92: Hemşire olmayı isterdim
93: Yükselmek için bir çok kimse yalan söylemekten çekinmez
94: Sonradan pişman olacağım pek çok şeyi yaptığım olur
95: Namazımı hemen hemen muntazaman kılarım
96: Ailemle pek az kavga ederim
97: Bazen zararlı yada çok kötü işler yapmam için içimde çok güçlü bir istek duyarım
98: Kıyamet gününe inanıyorum
99: Gürültülü eğlencelere katılmaktan hoşlanırım
100: Bildiğim bir konuda bir kimse saçma sapan ya da cahilce konuşursa onu hemen düzeltirim
101: Bence cinsel yönden kadınlar da erkekler kadar serbest olmalıdır,
102: En büyük mücadelemi kendimle yaparım
103: Vücudumda pek az seğirme ve kasılma olur
104: Başıma ne gelirse gelsin aldırış etmiyorum
105: Keyfim yerinde olmadığı zaman tersliğim üzerimdedir
106: Çoğu zaman büyük bir hata ya da kötülük yaptığım korkusuna kapılırım
107: Çoğu zaman mutluyumdur
108: Çoğu zaman bana kafam şişmiş ya da burnum tıkanmış gibi gelir
109: Bazı kimseler o kadar amirane davranırlar ki haklı bile olsalar içimden dediklerinin aksini yapmak gelir
110: Bana kötülük etmek isteyen biri var
111: Sırf heyecanlanmak için tehlikeli bir işe girişmedim
112: Doğru bildiğim şeyler için çoğu zaman direnmek zorunda kalırım
113: Kanunların uygulanması gerektiğine inanırım
114: Çoğu zaman başım sıkı bir çember içindeymiş gibi hissederim
115: Ahirete inanırım
116: Bahse girdiğim yarış ya da oyunlardan daha çok zevk alırım
117: Bir çok kimseler daha çok yakalanmaktan korktukları için dürüsttürler
118: Dersten kaçtığım için ara sıra müdüre gönderildiğim oldu
119: Konuşma tarzım her zamanki gibidir (Daha yavaş ya da hızlı değil, yayvanlaşmış ya da kısık değil)
120: Evde sofra adabına dışarıdaki kadar dikkat etmem
121: Aleyhimde bazı tertipler kurulduğuna inanıyorum
122: Tanıdığım insanların çoğu kadar becerikli ve zeki olduğuma sanıyorum
123: Beni takip edenler olduğuna inanıyorum
124: Bir çokları kaybetmektense çıkarlarını korumak için pek doğru olmayan yollara başvururlar
125: Midemden oldukça rahatsızım
126: Tiyatrodan hoşlanırım
127: Dertlerimin çoğundan kimin sorumlu olduğunu biliyorum
128: Kan görünce korkmam ya da fenalaşmam
129: Bazen ters ve suratsız olurum
130: Hiç bir zaman kan kusmadım ya da kan tükürmedim
131: Hastalığa yakalanacağım diye kaygılanmam
132: Çiçek koleksiyonu yapmayı ve evde çiçek yetiştirmeyi severim
133: Hiç bir zaman normal olmayan cinsel ilişkilere girişmedim
134: Bazen kafamdaki düşünceler o kadar hızlıdır ki söylemeyi yetiştiremem
135: Fark edilmeyeceğimden emin olsam sinemaya biletsiz girerdim
136: Bana iyilik yapan kimsenin genel olarak gizli bir amacı olabileceğini düşünürüm
137: Aile hayatımın tanıdığım kimselerin çoğununki kadar iyi olduğuna inanıyorum
138: Eleştiri beni çok kırar
139: Bazen sanki kendimi ya da başkasını incitmek zorundaymışım gibi hissederim
140: Yemek pişirmeyi severim
141: Davranışlarımı çoğu zaman etrafımdakilere göre ayarlarım
142: Bazen hiçbir işe yaramadığımı düşünürüm
143: Çocukken başlarına ne gelirse gelsin aralarındaki birliği koruyan bir gruptaydım
144: Asker olmak isterim
145: Bazen biriyle yumruk yumruğa kavgaya girişmek istediğim olur
146: Seyahat edip gezip tozmadıkça mutlu olmam
147: Çabuk karar vermediğim için çok fırsat kaçırdım
148: Önemli bir iş üzerinde çalışırken başkalarının işimi yarıda kesmeleri sabrımı taşırır
149: Hatıra defteri tutardım
150: Oyunda kaybetmektense kazanmayı isterim
151: Biri beni zehirlemeye çalışıyor
152: Çoğu geceler zihnimi hiçbir şey kurcalamadan uykuya dalarım
153: Son birkaç yıl içinde sağlığım çoğu zaman iyiydi
154: Hiç sinir nöbeti ya da havale geçirmedim
155: Ne şişmanlıyorum ne de zayıflıyorum
156: Bir şeyler yapıp sonra ne yaptığımı hatırlayamadığım zamanlar oldu
157: Çoğu kez sebepsiz yere cezalandırıldım
158: Çabuk ağlarım
159: ŸOkuduğumu eskisi kadar iyi anlayamıyorum
160: Hayatımda hiç bir zaman kendimi şimdiki kadar iyi hissetmedim
161: Bazen başımda bir sizi hissederim
162: Birisinin bana kurnazca oyun etmesine çok içerlerim
163: Çabucak yorulmam
164: Üzerinde çalıştığım konularda okumayı ve incelemelerde bulunmayı severim
165: Önemli kimseleri tanımayı severim, çünkü böylece kendimi de önemli bir kimse gibi görürüm
166: Yüksek bir yerden aşağıya bakmaya korkarım
167: Ailemden herhangi birinin mahkemelik olması beni rahatsız etmez
168: Zihnimde bir gariplik var
169: Parayı ellemekten korkmam
170: Başkalarının hakkımda ne düşündükleri beni rahatsız etmez
171: Bir eğlencede başkaları yapsalar bile, ben taşkınlık yapmaktan rahatsız olurum
172: Çoğu kez utangaçlığımı örtbas etmek ihtiyacını duyarım
173: Okulu severdim
174: Hiç bayılma nöbeti geçirmedim
175: Pek az başım döner ya da hiç dönmez
176: Yılandan büyük bir korkum yoktur
177: Annem iyi bir kadındır
178: Hafızam genellikle iyidir
179: Cinsel konularda sıkıntım vardır
180: Yeni tanıştığım kimselerle konuşma konusu bulmada güçlük çekerim
181: Canım sıkıldıkça heyecan yaratmayı severim
182: Aklımı oynatmaktan korkuyorum
183: Dilencilere para vermeyi doğru bulmam
184: Sık sık nereden geldiğini bilmediğim sesler duyarım
185: Herkes kadar iyi işitirim
186: Bir şeyler yapmağa girişince ellerimin çok defa titrediğini fark ederim
187: Ellerimde beceriksizlik ya da sakarlık yok
188: Gözlerim yorulmadan uzun süre okuyabilirim
189: Çoğu zaman bütün vücudumda bir halsizlik duyarım
190: Başım pek az ağrır
191: Bazen utanınca çok terlerim
192: Yürürken dengemi hemen hemen hiç kaybetmem
193: Saman nezlesi ya da astım nöbetlerim yoktur
194: Hareketlerimi ve konuşmamı kontrol edemediğim fakat etrafımda olup bitenden haberdar olduğum nöbetler geçirdiğim oldu
195: Tanıdığım herkesi sevmem
196: Hiç görmediğim yerlere gitmekten hoşlanırım
197: Biri beni soymaya her şeyimi almaya çalışıyor
198: Çok az hayal kurarım
199: Çocuklara cinsiyetle ilgili temel gerçekler öğretilmelidir
200: Fikir ve düşüncelerimi çalmak isteyen biri var
201: Keşke bu kadar utangaç olmasam
202: Kendimi cezayı hak etmiş suçlu bir insan olarak görüyorum
203: Gazeteci olmak isterdim
204: Gazeteci olmasaydım daha çok tiyatro haberleri yazmaktan hoşlanırdım
205: Bazen çalmaktan ya da dükkanlardan eşya aşırmaktan kendimi alamam
206: Bir çok kimseden daha dindarımdır
207: Çeşitli oyun ve eğlencelerden hoşlanırım
208: Flört etmeyi severim
209: Günahlarımın affedilmeyeceğine inanıyorum
210: Her şeyin tanı aynı geliyor
211: Gündüzleri uyuyabilirim fakat gündüzleri uyuyamam
212: Evdekiler bana çocuk muamelesi yapıyor
213: Yürürken kaldırımdaki yarıklara basmamaya dikkat ederim
214: Cildimde üzülmeye değer kabarıklık ya da sivilce yok
215: Çok içki kullandım
216: Başka ailelere göre bizim evde sevgi ve arkadaşlık pek azdır
217: Sık sık kendime bir şeyleri dert edinirim
218: Hayvanların eziyet çektiğini görmek beni üzmez
219: İnşaat mütahitliğinden hoşlanacağımı sanıyorum
220: Annemi çok severim
221: Bilimden hoşlanırım
222: Karşılığını veremeyeceğim durumlarda bile arkadaşlarımdan yardım istemekte güçlük çekmem
223: Avlanmayı çok severim
224: Annem babam hep beraber olduğum kimselerden çok defa hoşlanmıyorlar
225: Bazen biraz dedikodu yaptığım olur
226: Ailemdeki bazı kimselerde canımı çok sıkan alışkanlıklar var
227: Uykuda gezdiğimi söylerler
228: Bazen alışılmamış bir kolaylıkla karar verebileceğimi hissediyorum
229: Çeşitli klüp ve derneklere üye olmak isterim
230: Kalbimin hızlı çarptığını hemen hemen hiç hissetmem ve çok seyrek nefesim tıkanır
231: Cinsiyet hakkında konuşmayı severim
232: Bazen üzerime çok fazla iş alırım
233: Pek çok insan karşı çıksa da kendi fikrimi sonuna kadar savunurum
234: Çabuk kızar ve çabuk unuturum
235: Aile kurallarından oldukça bağımsız ve özgürüm
236: Sıklıkla kara kara düşünürüm
237: Akrabalarımın hemen hepsi bana karşı anlayış gösterir
238: Zaman zaman perimde duramayacak huzursuzluk duyduğum devreler olur
239: Aşkta hayal kırıklığına uğradım
240: Görünüşüme hiç aldırmam
241: Kendi içimde tutup başkalarına söylenemeyen şeyler hakkında sık sık rüya görürüm
242: Bir çoklarından daha sinirli sayılmam
243: Hemen hemen hiç bir ağrı ve sızım yok
244: Davranışlarım başkalarınca yanlış anlaşılmaya elverişlidir
245: Ailem beni olduğumdan daha hatalı bulur
246: Boynumda sık sık kırmızı lekeler olur
247: Kimseden sevgi görmüyorum
248: Bazen ortada hiç bir neden yokken hatta işler kötüye gittiği zaman bile kendimi fazlasıyla mutlu hissederim
249: Öbür dünyada şeytan ve cehennem olduğuna inanırım
250: Hayatta önüne her geleni kapmağa çalışan insanları suçlamam
251: Kendimi kaybedip yaptığım işi aksattığım ve etrafımda olup bitenlerin farkında olmadığım zamanlar oldu
252: Hiç kimse başkasının derdine aldırış etmiyor
253: Hatalı davranışlarını görsem bile insanlara arkadaşça davranabilirim
254: Birbiriyle şakalaşan kimseler arasında olmayı severim
255: Seçimlerde bazen oyumu pek az tanıdığım kimselere veririm
256: Gazetelerin ilgi çeken tek yeri resimli mizah sayfasıdır
257: Yaptığım işlerde genel olarak başarı elde edeceğime inanırım
258: Allah’ın varlığına inanırım
259: İşe başlamada güçlük çekerim
260: øOkulda iken ağır öğrenenlerden biri idim
261: Ressam olsaydım çiçek resimleri yapardım
262: Daha güzel olmamam beni rahatsız etmez
263: Soğuk günlerde bile kolayca terlerim
264: Kendime tam anlamıyla güvenim vardır
265: Hiç kimseye güvenmemek en doğrusudur
266: Haftada bir ya da sık, çok heyecanlanırım
267: Topluluk içinde olduğumda üzerinde konuşacak uygun konular bulmada güçlük çekerim
268: Karamsar olduğum zaman heyecanlı bir olay hemen beni bu durumdan çıkarır
269: Bazen zevk için başkalarını kendimden korkuturum
270: Evden çıkarken kapının kilitli ve pencerenin kapalı olup olmadığı aklıma takılmaz
271: Başkalarının saflığını kendi çıkarlarında kullanan kimseleri ayıplamam
272: Bazen kendimi enerji dolu hissederim
273: Derimin bazı yerlerinde uyuşukluk hissederim
274: Görme gücüm eskisi kadar kuvvetlidir
275: Birisi zihnimi kontrol ediyor
276: Çocukları severim
277: Bazen bir madrabazın kurnazlığı beni o kadar eğlendirir ki, yakayı ele vermemesini dilerim
278: Çok defa tanımadığım kimselerin bana eleştirici gözle baktıklarını hissederim
279: Her gün gereğinden fazla su içerim
280: Bir çok kimseler kendilerine yararı dokunacağı için arkadaş edinirler
281: Kulaklarım çok az çınlar ya da uğuldar
282: Genellikle sevdiğim aile üyelerine karşı bazen nefret duyarım
283: Gazete muhabiri olsaydım en çok spor haberleri yazmayı isterdim
284: Hakkımda çok konuşulduğumdan eminim
285: Ara sıra açık saçık bir fıkraya güldüğüm olur
286: En çok yalnız olduğum zaman mutlu olurum
287: Arkadaşlarıma kıyasla beni korkutan şeyler çok azdır
288: Tekrarlanan mide bulantısı ve kusmalar bana sıkıntı verir
289: Bir suçlu avukatının becerikliliği sayesinde cezadan kurtulunca kanunlara karşı daima nefret duyarım
290: Çok gergin bir hava içinde çalışıyorum
291: Hayatımda bir ya da birkaç kere birisinin beni hipnotize ederek bana bir şeyler yaptığını hissettim
292: Başkaları benimle konuşuncaya kadar ben onlarla konuşmaya başlamam
293: Birisi zihnimi etkilemeye çalışıyor
294: Kanunla hiç başım derde girmedi
295: Masal okumayı severim
296: Hiçbir neden yokken kendimi son derecede neşeli hissettiğim zamanlar olur
297: Cinsiyetle ilgili düşünceler beni rahatsız eder
298: Birkaç kişinin birlikte başları derde girince en iyisi yakalarını kurtarmak için aynı hikayeyi uydurmak ve bundan caymamaktır
299: Duygularımın birçok kimselerden yoğun olduğunu düşünürüm
300: Hayatımda hiçbir zaman bebek oynamaktan hoşlanmadım
301: Çoğu zaman hayat benim için bir yüktür
302: Cinsel davranışlarımdan dolayı hiçbir zaman başım derde girmedi
303: Bazı konularda o kadar alınganım ki onlar hakkında konuşmam bile
304: Okulda sınıf karşısında konuşma bana çok güç gelirdi
305: Başkalarıyla beraber olduğum zaman bile kendimi yalnız hissederim
306: Bana karşı mümkün olan anlayış gösteriliyor
307: İyi beceremediğim oyunları oynamağa yanaşmam
308: Zaman zaman evi bırakıp gitmeyi çok istemişimdir
309: Birçokları kadar çabuk arkadaş edinebildiğimi sanıyorum
310: Cinsel hayatım doyurucudur
311: Gençlik yıllarımda bir devre ufak tefek şeyler çaldım
312: İnsanların arasında olmaktan hiç hoşlanmam
313: Değerli eşyasını tedbirsizce ortada bırakıp çalınmasına neden olan kimse bunu çalan kadar hatalıdır
314: Ara sıra söylenemeyecek kadar kötü şeyler düşünürüm
315: Hayatın hep kötü tarafları bana nasip olmuştur
316: Hemen hemen herkesin başını derde sokmamak için yalan söyleyebileceğine inanırım
317: Birçok kimselerden daha hassasım
318: Günlük hayatım beni ilgilendiren şeylerle dolu
319: İnsanların çoğu başkalarına yardım etmek için zahmete girmekten hoşlanmazlar
320: Rüyalarımın çoğu cinsel konularla ilgilidir
321: Kolaylıkla mahcup olurum
322: Para ve işi kendime dert edinirim
323: Başımdan çok tuhaf ve acayip olaylar geçmiştir
324: Hiç kimseye aşık olmadım
325: Ailemin yaptığı bazı şeyler beni korkutmuştur
326: Bazen kontrol edemediğim gülme ve ağlama nöbetlerine tutulurum
327: Annem ya da babam çok defa beni makul bulmadığım emirlere bile itaat ettirdiler
328: Zihnimi bir konu ya da iş üzerinde toplamakta güçlük çekerim
329: Hemen hemen hiç rüya görmedim
330: Hiç felç geçirmedim ya da kaslarımda olağan üstü bir halsizlik duymadım
331: Eğer insanlar sırf düşmanlık olsun diye beni engellemeselerdi daha başarılı olurdum
332: Bazen nezle olmadığım halde sesim çıkmaz ya da değişir
333: Beni hiç kimse anlamıyor
334: Bazen tuhaf korkular duyarım
335: Zihnimi bir konu üzerinde toplayamam
336: İnsanlara karşı sabrım çabuk tükenir
337: Çoğunlukla bir takım şeyler ve kimseler için meraklanıp huzursuzlaşırım
338: Hayatımın çoğu kimselerinkinden daha fazla tasa ve kaygı içinde geçtiğine eminim
339: Çoğu zaman ölmüş olmayı isterdim
340: Bazen o kadar heyecanlanırım ki uykuya dalmam güçleşir
341: Bazen beni rahatsız edecek kadar iyi işitirim
342: Bana söyleneni hemen unuturum
343: Önemsiz ufak şeylerde bile karar verip işe girişmeden önce durur ve düşünürüm
344: Gördüğüm kimse ile karşılaşmamak için sıklıkla yolumu değiştiririm
345: Sıklıkla olup bitenler bana gerçek değilmiş gibi gelir
346: Reklamlardaki ampuller gibi önemsiz şeyleri sayma alışkanlığım vardır
347: Bana gerçekten kötülük yapmak isteyen hiç bir düşmanım yoktur
348: Bana umduğumdan fazla dostluk gösteren insanlara karşı tetikte bulunmağa çalışırım
349: Acayip ve tuhaf düşüncelerim vardır
350: Yalnızken garip şeyler duyarım
351: Küçük bir seyahat için bile evden ayrılırken telaşlanır ve kaygılanırım
352: Beni incitmeyeceğini bildiğim şeylerden ya da insanlardan bile korktuğum oldu
353: Başkalarının daha önce toplanıp konuştuğu odaya girmekten çekinmem
354: Bıçak gibi çok keskin ve sivri şeyler kullanmaktan korkarım
355: Sevdiğim kimseleri bazen incitmekten hoşlanırım
356: Dikkatimi bir konu üzerinde toplamada birçok kişiden daha fazla güçlük çekerim
357: Yeteneğimi küçümsediğim için birçok defalar başladığım işi yarıda bıraktım
358: Kötü ve çok defa korkunç kelimeler zihnimi kurcalar ve bunlardan kendimi kurtaramam
359: Bazen önemsiz düşünceler aklımdan geçer ve beni günlerce rahatsız eder
360: Hemen hemen her gün beni korkutan bir şey olur
361: Her şeyi kötüye yorma eğilimindeyim
362: Birçok kimselerden daha çok hassasım
363: Bazen sevdiğim kimselerin beni incitmesinden hoşlandığım oldu
364: Hakkımda onur kırıcı ve kötü sözler söylüyorlar
365: Kapalı yerlerde huzursuzluk duyarım
366: İnsanlar içinde bile olsam çok defa kendimi yalnız hissederim
367: Yangından korkmam
368: Sonradan pişman olacağım şeyler yapmak ya da söylemek korkusuyla bazen bir kimseden uzak durduğum oldu
369: Kararsızlığım yüzünden yapılması gerekli birçok işi yapamamışımdır
370: Çalışırken acele etmek zorunda olmaktan nefret ederim
371: Aşırı derecede kendini dinleyen bir insan değilim
372: Elimdeki işi en iyi şekilde yapmayı isterim
373: Yalnızca bir tek doğru din olduğundan eminim
374: Ara sıra zihnim her zamankinden daha ağır işler
375: Çok mutlu olduğum ve iyi çalıştığım zamanlarda neşesiz veya dertli bir insanla karşılaşmak keyfimi tamamen kaçırır
376: Polisler genellikle dürüsttür
377: Toplantılarda kalabalığa karışmaktan çok yalnız başıma oturur ya da bir tek kişiyle ahbaplık ederim
378: Kadınları sigara içerken görmekten hoşlanmam
379: Çok nadiren karamsarlığa kapılırım
380: Ne yapsam zevk alamıyorum
381: Kolay öfkelenen biri olduğumu söylerler
382: Yapmak istediğim şeylere karar verirken, başkalarının ne düşüneceğini dikkate almam
383: İnsanlar çoğu zaman beni hayal kırıklığına uğratırlar
384: Kendimle ilgili her şeyi anlatabileceğim hiç kimse yok
385: Şimşek çakması da korkularımdan biridir
386: Çok tertipli ve titizim
387: Ailem her davranışıma fazla karışıyor
388: Karanlıkta yalnız kalmaktan korkarım
389: Tasarlamış olduğum planlar çok defa o kadar güçlükle dolu göründü ki bunlardan vazgeçmek zorunda kaldım
390: Birinin hatasını önleme gayretimin yanlış anlaşılmasına çok üzülürüm
391: Dansa gitmeyi severim
392: Fırtınadan çok korkarım
393: Yük çekmeyen atlar dövülmeli ya da kamçılanmalıdır
394: Başkalarına sık sık akıl danışırım
395: Gelecek, bir insanın ciddi planlar yapamayacağı kadar belirsizdir
396: İşler yolunda gittiği zaman bile çoğu kez her şeye karşı bir aldırmazlık içinde olduğumu hissederim
397: Bazen güçlükler öylesine üst üste gelir ki onlarla baş edemeyecekmişim gibi hissederim
398: Çoğu kez keşke tekrar çocuk olsaydım diye düşünürüm
399: Kolay kolay kızmam
400: Eğer bana fırsat verilse dünya için çok yararlı işler yapabilirim
401: Sudan hiç korkmam
402: Ne yapacağıma karar vermeden önce uzun uzun düşünürüm
403: Birçok şeyin olup bittiği böyle bir devirde yaşamak hoş bir şey
404: Hatalarını düzelterek kendilerine yardım etmeye çalıştığım insanlar amacımı çoğu kez yanlış anlarlar
405: Yutkunmakta güçlük çekmem
406: Uzman dendiği halde benden pek fazla bilgili olmayan insanlarla sık sık karşılaşırım
407: Genel olarak sakinim ve kolay sinirlenmem
408: Bazı konular hakkında hislerimi o kadar gizleyebilirim ki insanlar bilmeden beni incitebilirler
409: Elimde olmadan çok ufak bir şeyden münakaşa çıkarıp karşımdakini kırıyorum
410: Madrabazı kendi silahı ile alt etmekten hoşlanırım
411: İyi tanıdığım bir kimsenin başarısını duyduğum zaman adeta kendimi başarısızlığı uğramış hissederim
412: Hastalandığım zaman doktora gitmekten korkmam
413: Günahlarım için ne kadar ağır ceza görsem iyidir
414: Hayal kırıklıklarını o kadar ciddiye alırım ki bunları zihnimden söküp atamam
415: Fırsat verilse iyi bir önder olurum
416: Yakınlarımın sağlığından çok endişe ederim
417: Sırada beklerken biri önüme geçmeye kalkışırsa ona çıkışırım
418: Bazen hiç bir işe yaramadığımı düşünürüm
419: Küçükken okuldan sık sık kaçardım
420: Başımdan dinle ilgili olağan üstü yaşantılar geçti
421: Ailemde çok sinirli insanlar var
422: Ailemde bazı kişilerin yapmış olduğu işler beni utandırmıştır
423: Balık tutmayı çok severim
424: Hemen hemen her zaman açlık duyarım
425: Sık sık rüya görürüm
426: Kaba ya da can sıkıcı insanlara karşı bazen sert davrandığım olur
427: Açık saçık hikayelerden utanıp rahatsız olurum
428: Gazetelerin baş yazılarını okumaktan hoşlanırım
429: Ciddi konular üzerinde verilen konferansları dinlemekten hoşlanırım
430: Karşı cinsten olanları çekici bulurum
431: Başa gelebilecek talihsizlikler beni oldukça telaşlandırır
432: Kuvvetli siyasi fikirlerim vardır
433: Bir zamanlar hayali arkadaşlarım vardı
434: Otomobil yarışçısı olmayı isterdim
435: Genel olarak kadınlarla çalışmayı tercih ederim
436: İnsanlar genel olarak başkalarının haklarına saygı göstermekten çok kendi haklarına saygı gösterilmesini isterler
437: Kanuna aykırı davranmadan kanunun bir gediğinden yararlanmakta zarar yoktur
438: Bazı insanlardan o kadar nefret ederim ki ettiklerini bulunca içimden oh derim
439: Beklemek zorunda kalmak beni sinirlendirir
440: Başkalarına anlatmak için hoş fıkraları hatırımda tutmaya çalışırım
441: Uzun boylu kadınlardan hoşlanırım
442: Üzüntü yüzünden uyuyamadığım zamanlar oldu
443: Başkalarının gereği gibi yapamadığımı sandığı şeyleri yapmaktan vaz geçtiğim oldu
444: Başkalarının cahilce inançlarının düzeltmeye çalışmam
445: Küçükken heyecan veren şeyler yapmaktan hoşlanırdım
446: Az parayla oynanan kumardan hoşlanırım
447: Mastürbasyonda kendi cinsimle ilgili hayal beni tahrik eder
448: Sokakta, otobüs ve dükkanlarda bana bakan insanlardan rahatsız olurum
449: İnsanlarla bir arada olmayı sağladığı için toplantı ve davetleri severim
450: Kalabalığın verdiği coşkudan hoşlanırım
451: Neşeli arkadaşlar arasına karışınca üzüntülerimi unuturum
452: Arkadaş edinemiyorum
453: Küçükken mahalledeki arkadaş ya da akran gruplarına katılmaktan hoşlanmazdım
454: Orman ya da dağdaki bir kulübede tek başıma yaşamaktan mutlu olabilirim
455: İçinde bulunduğum grubun dedikodularına ve konuşmalarına sıklıkla konu olmam
456: İnsan makul bulmadığı kanunlara aykırı hareketlerinden ötürü cezalandırılmamalıdır
457: Bence insan hiç bir zaman alkollü içkiyi ağzına almamalıdır
458: Çocukken benimle en fazla ilgilenen erkek baba, üvey baba vb bana karşı çok sert davranırdı
459: Çaba göstermekle yenemeyeceğimi bildiğim bazı kötü alışkanlıklarım var
460: Az içki kullandım ya da hiç kullanmadım
461: Kısa bir zaman için bile olsa başladığım işi bir kenara bırakmak bana güç gelir
462: Küçük abdestimi yapmada ya da tutmada güçlük çekmem
463: Sek sek oyunu oynamaktan hoşlanırdım
464: Hiç hayal görmedim
465: Bir kaç kez hayatım boyunca yaptığım işte hevesimi yitirdiğim olmuştur
466: Doktor önerisi dışında hiçbir ilaç ya da uyku hapı kullanmadım
467: Çok defa otomobil, plaka numarası gibi hiç önemli olmayan numaraları ezberlerim
468: Sıklıkla sinirli ve asık suratlı olurum
469: Onlardan önce düşündüğüm için başkaları benim fikirlerimi kıskanıyorlar
470: Cinsiyetle ilgili şeylerden nefret ederim
471: Okulda hal ve gidişten kırık not alırdım
472: Yangın karşısında büyülenmiş gibi olurum
473: Mümkün olduğu kadar kalabalıktan uzak kalmaya çalışırım
474: Başkalarından daha sık küçük abdeste çıkmam
475: Sıkıştırıldığım zaman gerçeğin ancak bana zarar vermeyecek kısmını söylerim
476: Tanrı bana özel bir görev vermiştir
477: Arkadaşlarımla birlikte işlediğim bir suçtan eşit şekilde suçlu olduğum zaman onları ele vermektense bütün suçu üzerime almayı tercih ederim
478: Çok değişik bir aile ortamından gelmiş olmayı isterdim
479: Yabancılarla tanışmaktan kaçınmam
480: Karanlıktan çok defa korkarım
481: Bir şeyden kurtulmak için hasta numarası yaptığım olmuştur
482: Trende, otobüste vb rastladığım kimselerle çok defa konuşurum
483: Peygamberimiz göğe çıkma gibi mucizeler göstermiştir
484: Homoseksüelliği çok iğrenç buluyorum
485: Bir erkek bir kadınla beraber olunca genel olarak onun cinsiyetiyle ilgili şeyler düşünür
486: İdrarımda hiç bir zaman kan görmedim
487: Uğraştığım iş yolunda gitmeyince hemen vaz geçerim
488: Sık sık dua ederim
489: Yaşamı yalnızca üzüntülü, sıkıntılı tarafları ile benimseyen insanlara sempati duyarım
490: Haftada birkaç kere kuran okurum
491: Sadece bir tek dinin doğruluğuna inananlara tahammül edemem
492: Zelzele düşüncesi beni çok korkutur
493: Tam dikkat isteyen işleri, beni dikkatsizliğe sürükleyen işlere tercih ederim
494: Kapalı ve küçük yerlerde bulunmaktan çok rahatsız olurum
495: Kusurlarını düzeltmeye çalıştığım insanlarla genel olarak gayet açık konuşurum
496: Eşyayı hiçbir zaman çift görmem (Yani tek olan şeyleri çift görmem)
497: Macera hikayelerinden hoşlanırım
498: Açık sözlü olmak her zaman iyidir
499: Gerçekten önemsiz olan bir şey üzerinde bazen sebepsiz olarak haddinden fazla üzüldüğüm olur
500: Bana parlak gelen bir fikre hemen kapılır giderim
501: Başkalarından yardım beklemektense genel olarak bir işi kendi başıma yapmayı tercih ederim
502: Herhangi bir olay hakkındaki görüşümü başkalarına açıkça belirtmekten hoşlanırım
503: Başkalarının hareketlerin çok beğenip beğenmediğimi pek belli etmem
504: Değersiz gördüğüm ya da acıdığım kimseye bu duygularımı belli etmekten çekinmem
505: Zaman zaman kendimi öyle güçlü ve enerjik hissederim ki böyle zamanlarda günlerce uykuya ihtiyaç duymadığım olur
506: Sinirleri çok gergin bir insanım
507: İşler iyi gidince aslan payını kendilerine alan fakat hata yapılanca bunu başkalarının üzerine atan insanlarla karşılaştım
508: Koku alma duyum herkes kadar iyidir
509: Bazen çekingenliğim yüzünden hakkımı arayamam
510: Pislik ve kir beni ürkütüp iğrendirir
511: Herkesten gizli tuttuğum bir hayal dünyam var
512: Yıkanmaktan hoşlanmam
513: Kış mevsimini severim
514: Erkek gibi davranan kadınlardan hoşlanırım
515: Evimizde daima gerekli ihtiyaç maddeleri bulunurdu (Yeteri kadar yiyecek, giyecek vb gibi)
516: Ailemde çabuk kızan kimseler var
517: Hiç bir şeyi iyi yapamam
518: Bazı durumlarda olduğumdan daha fazla üzüntülü görünmeye çalıştığım olmuştur
519: Cinsel organlarımda bir bozukluk var
520: Genel olarak görüşlerimi kuvvetle savunurum
521: Bir grup içinde konuşma yapmam ve çok iyi bildiğim bir konuda fikrimi söylemem istenince kaygılanmam
522: Örümcekten korkmam
523: Yüzüm hemen hemen hiç kızarmaz
524: Kapı tokmaklarından hastalık veya mikrop alacağımdan korkmam
525: Bazı hayvanlardan ürkerim
526: Gelecek bana ümitsiz görünüyor
527: Ailem ve yakın akrabalarım birbirleriyle oldukça iyi geçinirler
528: Yüzüm başkalarından daha sik kızarmaz
529: Pahalı elbiseler giymeyi isterim
530: Sebepsiz yere sık sık içim sıkılıyor ve ağlamak istiyorum
531: Bir konu üzerinde karar verdiğimi zannetsem bile başka biri fikrimi kolayca değiştirebilir
532: Acıya başkaları kadar ben de dayanabilirim
533: Sık sık geğirmekten şikayetim yoktur
534: Çoğunlukla başladığım işten en son vazgeçen ben olurum
535: Hemen hemen her zaman ağzımda kuruluk olur
536: Beni acele ettirenlere kızarım
537: Afrika'da aslan avına çıkmak isterdim
538: Terzilikten hoşlanabileceğimi sanıyorum
539: Fareden korkmam
540: Yüzüme hiç felç inmedi
541: Cildime ufak bir şeyin dokunmasından çok huylanırım
542: Şimdiye kadar rengi kapkara büyük abdest yapmadım
543: Haftada birkaç kez korkunç bir şey olacakmış duygusuna kapılırım
544: Çoğu zaman yorgunluk hissederim
545: Bazen aynı rüyayı tekrar tekrar görürüm
546: Tarih okumaktan hoşlanırım
547: Toplantı ve kalabalık eğlencelerden hoşlanırım
548: Elimdeyse açık saçık numaraların yapılacağı eğence yerlerine gitmem
549: Karşıma çıkacak güçlüklerden korkak ve kaçarım
550: Kapı mandallarını onarmaktan hoşlanırım
551: Bazen başkalarının kafamın içindekilerindi okuduğundan eminim
552: Bilimsel yayınları okumaktan hoşlanırım
553: Açık yerlerde veya geniş meydanlarda tek başıma kalmaktan korkarım
554: Sıkıntım oldukça alkol alırım
555: Bazen çıldıracakmış gibi olurum
556: Kılık kıyafetime çok itina ederim
557: Hayatı fazla ciddiye almıyorum
558: Birçok kimseler kötü cinsel faaliyetlerinden dolayı suçludurlar
559: Gece yarısı çoğunlukla korkuya kapıldığım olur
560: Bir şeyi nereye koyduğumu unutmaktan çok şikayetçiyimdir
561: Ailem benim için büyük bir dayanaktır
562: Çocukken en fazla bağlandığım ve hayran kaldığım kimse bir kadındı
563: Macera hikayelerini aşk hikayelerinden daha çok severim
564: Yapmak istediğim fakat başkalarının beğenmediği bir işten kolayca vazgeçerim
565: Yüksek bir yerde iken içimden atlama isteği gelir
566: Sinemalardaki aşk sahnelerini severim
}


# Used to store the original results of the questionnaire
Ans = {

}


def start():
    """
    Guide section, including abstract and instruction

    :return: None
    """
    print(Ins1)
    time.sleep(5)

    while 1:
        print('The quiz will start next, do you want to start？')
        print('1. yes    0. no')
        go = input('> ')
        if go == '1':
            print('-' * 65)
            print('The quiz will officially start in 30 seconds, please read and understand the following content carefully')
            time.sleep(3)
            print(Ins2)
            time.sleep(27)
            break
        elif go == '0':
            print('Thanks for using this program, bye！')
            time.sleep(3)
            exit(0)
        else:
            print('Input error, please re-enter according to the test requirements！')
            continue


def answer():
    """for debug
    random answer

    :return: '0' or '1'
    :rtype: str
    """
    ans = randint(0, 1)
    return str(ans)


def test():
    """
    test section

    :return: None
    """
    global Sex
    global Age

    print('The quiz has officially started！')
    print('-' * 65)
    time.sleep(3)

    while 1:
        print('x1. my gender is')
        print('1. male    0.female')
        Sex = input('> ')
        if Sex == '1' or Sex == '0':
            break
        else:
            print('Input error, please re-enter according to the test requirements！')
            continue

    while 1:
        print('-' * 65)
        print('x2. Please enter your age')
        Age = input('> ')

        if str.isdigit(Age):
            if 13 <= int(Age) <= 70:
                break
            else:
                print('This quiz is not intended for this age range, thanks for using！')
                time.sleep(3)
                exit(0)
        else:
            print('Input error, please re-enter according to the test requirements！')
            continue

    for i in range(len(Que)+1):
        if i == 73:
            if Sex == '1':
                temp_que = str(i+1) + '. ' + (Que[i+1][Que[i+1].find('m')+1: Que[i+1].find('f')])
            else:
                temp_que = str(i+1) + '. ' + (Que[i+1][Que[i+1].find('f') + 1:])
        elif i == len(Que):
            temp_que = str(len(Que)+1) + '.' + 'I promise to complete this quiz seriously and honestly under the guidance of professionals'
        else:
            temp_que = str(i+1) + '. ' + Que[i+1]

        while 1:
            print('-' * 65)
            print(temp_que)
            print('1. yes    0. no')
            temp_ans = input('> ')
            # temp_ans = answer()  # for debug
            # print('> ' + str(temp_ans))  # for debug
            if temp_ans == '1' or temp_ans == '0':
                Ans[i+1] = temp_ans
                # print(Ans)  # for debug
                break
            elif temp_ans == 'bomb':    # for debug
                exit(0)
            else:
                print('Input error, please re-enter according to the test requirements！')
                continue

    print('-' * 65)
    print('The quiz is over, thank you for your cooperation！')
    print('-' * 65)


def is_diff(a, b):
    """
    Scoring for difference
    Add point if different

    :param a: the first para
    :param b: the second para
    :return: 0 or 1

    :type a: str
    :type b: str
    :rtype: int
    """
    if a != b:
        return 1
    else:
        return 0


def is_true(t):
    """
    Positive scoring
    Add point if True

    :param t: the para under test
    :return: 0 or 1

    :type t: str
    :rtype: int
    """
    if t == '1':
        return 1
    else:
        return 0


def is_false(t):
    """
    reverse scoring
    Add point if False

    :param t: the para under test
    :return: 0 or 1

    :type t: str
    :rtype: 0 or 1
    """
    if t == '0':
        return 1
    else:
        return 0


def norm_select(sex):
    """
    Select Norm Table (China 1982 Edition)
    the norm select (based on Chinese 1982's)

    :param sex: the subjects' sex
    :return: Norm_M, Norm_SD

    :type sex: str
    :rtype: None
    """
    global Norm_M
    global Norm_SD

    # male pattern
    # male norm
    if sex == '1':
        Norm_M = {
            'L': 5.70,
            'F': 13.68,
            'K': 13.00,
            'Hs': 8.78,
            'D': 26.16,
            'Hy': 22.07,
            'Pd': 18.98,
            'Mf': 27.56,
            'Pa': 12.84,
            'Pt': 17.86,
            'Sc': 23.01,
            'Ma': 18.48,
            'Si': 34.51,
            'Hs+0.5K': 15.42,
            'Pd+0.4K': 24.38,
            'Pt+1.0K': 31.14,
            'Sc+1.0K': 36.47,
            'Ma+0.2K': 21.22,
            'Mas': 18.86,
            'Dy': 26.09,
            'Do': 15.39,
            'Re': 20.54,
            'Cn': 25.26
        }
        Norm_SD = {
            'L': 2.52,
            'F': 6.86,
            'K': 4.66,
            'Hs': 4.75,
            'D': 4.97,
            'Hy': 5.36,
            'Pd': 4.36,
            'Mf': 4.04,
            'Pa': 3.92,
            'Pt': 7.93,
            'Sc': 10.15,
            'Ma': 5.26,
            'Si': 6.88,
            'Hs+0.5K': 4.79,
            'Pd+0.4K': 4.27,
            'Pt+1.0K': 5.71,
            'Sc+1.0K': 8.24,
            'Ma+0.2K': 4.88,
            'Mas': 7.45,
            'Dy': 8.05,
            'Do': 3.12,
            'Re': 4.13,
            'Cn': 3.76
        }
    # female pattern
    # female norm
    else:
        Norm_M = {
            'L': 5.64,
            'F': 11.69,
            'K': 12.25,
            'Hs': 9.83,
            'D': 28.40,
            'Hy': 22.82,
            'Pd': 18.29,
            'Mf': 31.83,
            'Pa': 12.62,
            'Pt': 18.77,
            'Sc': 22.50,
            'Ma': 16.64,
            'Si': 37.27,
            'Hs+0.5K': 16.35,
            'Pd+0.4K': 23.33,
            'Pt+1.0K': 31.17,
            'Sc+1.0K': 34.89,
            'Ma+0.2K': 19.18,
            'Mas': 20.43,
            'Dy': 29.12,
            'Do': 15.10,
            'Re': 21.78,
            'Cn': 24.86
        }
        Norm_SD = {
            'L': 2.48,
            'F': 5.02,
            'K': 4.26,
            'Hs': 4.98,
            'D': 5.04,
            'Hy': 5.54,
            'Pd': 4.45,
            'Mf': 3.86,
            'Pa': 3.93,
            'Pt': 7.82,
            'Sc': 9.57,
            'Ma': 5.16,
            'Si': 6.71,
            'Hs+0.5K': 4.95,
            'Pd+0.4K': 4.44,
            'Pt+1.0K': 5.86,
            'Sc+1.0K': 7.63,
            'Ma+0.2K': 4.89,
            'Mas': 7.35,
            'Dy': 7.61,
            'Do': 2.76,
            'Re': 3.13,
            'Cn': 3.70
        }



def trans_t(score, m, sd):
    """
    Standard T-score calculation formula
    Standard T point conversion formula

    :param score: original score
    :param m: normative mean value
    :param sd: normative standard deviation
    :return: standard T score

    :type score: int
    :type m: float
    :type sd: float
    :rtype: int
    """
    t = round(50 + 10*(score - m)/sd)
    return t


def scale_q(ori_score=0, pro_score=0):
    """
    Validity Scale - Question Score Q
    the score of Q (? or question) scale,

    Since the subjects are not allowed to have empty questions, only the contradictory numbers of the 16 repeated questions are recorded
    because subjects were not allowed to have blank questions,
    just record the number of contradictions of 16 repeated questions

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    # raw score
    # original score
    temp1 = [8, 13, 15, 16, 20, 21, 22, 23, 24, 32, 33, 35, 37, 38, 305, 317]
    temp2 = [318, 290, 314, 315, 310, 308, 326, 288, 333, 328, 323, 331, 302, 311, 366, 362]

    for i in range(len(temp1)):
        ori_score += is_diff(Ans[temp1[i]], Ans[temp2[i]])

    temp = ori_score
    pro_score += temp

    return ori_score, pro_score


def scale_l(ori_score=0, pro_score=0):
    """
    Validity Scale - Lying Score L
    the score of L (lie) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    # raw score
    # original score
    temp = [15, 30, 45, 60, 75, 90, 105, 120, 135, 150, 165, 195, 225, 255, 285]

    for i in temp:
        ori_score += is_false(Ans[i])

    pro_score += trans_t(ori_score, Norm_M['L'], Norm_SD['L'])

    return ori_score, pro_score


def scale_f(ori_score=0, pro_score=0):
    """
    Validity Scale - Masquerade Score F
    the score of F (infrequency or fake bad) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    # raw score
    # original score
    temp_t = [14, 27, 31, 34, 35, 40, 42, 48, 49, 50, 53, 56, 66, 85, 121, 123, 139, 146, 151, 156, 168, 184, 197, 200,
              202, 205, 206, 209, 210, 211, 215, 218, 227, 245, 246, 247, 252, 256, 269, 275, 286, 288, 291, 293]
    temp_f = [17, 20, 54, 65, 75, 83, 112, 113, 115, 164, 169, 177, 185, 196, 199, 220, 257, 258, 272, 276]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['F'], Norm_SD['F'])

    return ori_score, pro_score


def scale_k(ori_score=0, pro_score=0):
    """
    Validity Scale-Adjusted Score K
    the score of K (defensiveness) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [96]
    temp_f = [30, 39, 71, 89, 124, 129, 134, 138, 142, 148, 160, 170, 171, 180, 183, 217, 234, 267, 272, 296, 316, 322,
              368, 370, 372, 373, 375, 386, 394]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['K'], Norm_SD['K'])

    return ori_score, pro_score


def scale_hs(ori_score=0, pro_score=0, pro_score_add_k=0):
    """
    Clinical Scale-1 Hypochondria Hs
    the score of Hs (hypochondriasis) scale

    :param ori_score: original score
    :param pro_score: processing score
    :param pro_score_add_k: processing score added 0.5K
    :return: ori_score, pro_score, pro_score_add_k

    :rtype: int, int, int
    """
    temp_t = [23, 29, 43, 62, 72, 108, 114, 125, 161, 189, 273]
    temp_f = [2, 3, 7, 9, 18, 51, 55, 63, 68, 103, 130, 153, 155, 163, 175, 188, 190, 192, 230, 243, 274, 281]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    k, ignore = scale_k()
    pro_score += trans_t(ori_score, Norm_M['Hs'], Norm_SD['Hs'])
    pro_score_add_k += trans_t(ori_score + round(0.5 * k), Norm_M['Hs+0.5K'], Norm_SD['Hs+0.5K'])

    return ori_score, pro_score, pro_score_add_k


def scale_d(ori_score=0, pro_score=0):
    """
    Clinical Scale-2 Depression D
    the score of D (depression) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [5, 32, 41, 43, 52, 67, 86, 104, 130, 138, 142, 158, 159, 182, 189, 193, 236, 259, 288, 290]
    temp_f = [2, 8, 9, 18, 30, 36, 39, 46, 51, 57, 58, 64, 80, 88, 89, 95, 98, 107, 122, 131, 145, 152, 153, 154, 155,
              160, 178, 191, 207, 208, 233, 241, 242, 248, 263, 270, 271, 272, 285, 296]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['D'], Norm_SD['D'])

    return ori_score, pro_score


def scale_hy(ori_score=0, pro_score=0):
    """
    Clinical Scale-3 Hysteria Hy
    the score of Hy (hysteria) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [10, 23, 32, 43, 44, 47, 76, 114, 179, 186, 189, 238, 253]
    temp_f = [2, 3, 6, 7, 8, 9, 12, 26, 30, 51, 55, 71, 89, 93, 103, 107, 109, 124, 128, 129, 136, 137, 141, 147, 153,
              160, 162, 163, 170, 172, 174, 175, 180, 188, 190, 192, 201, 213, 230, 234, 243, 265, 267, 274, 279, 289,
              292]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Hy'], Norm_SD['Hy'])

    return ori_score, pro_score


def scale_pd(ori_score=0, pro_score=0, pro_score_add_k=0):
    """
    Clinical Scale-4 Psychopathy Pd
    the score of Pd (psychopathic deviate) scale

    :param ori_score: original score
    :param pro_score: processing score
    :param pro_score_add_k: processing score added 0.4K
    :return: ori_score, pro_score, pro_score_add_k

    :rtype: int, int, int
    """
    temp_t = [16, 21, 24, 32, 33, 35, 38, 42, 61, 67, 84, 94, 102, 106, 110, 118, 127, 215, 216, 224, 239, 244, 245,
              284]
    temp_f = [8, 20, 37, 82, 91, 96, 107, 134, 137, 141, 155, 170, 171, 173, 180, 183, 201, 231, 235, 237, 248, 267,
              287, 289, 294, 296]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    k, ignore = scale_k()
    pro_score += trans_t(ori_score, Norm_M['Pd'], Norm_SD['Pd'])
    pro_score_add_k += trans_t(ori_score + round(0.4 * k), Norm_M['Pd+0.4K'], Norm_SD['Pd+0.4K'])

    return ori_score, pro_score, pro_score_add_k


def scale_mf(ori_score=0, pro_score=0):
    """
    Clinical Scale-5 Manliness/Femininity Mf
    the score of Mf (masculinity-femininity) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    # masculine feminization
    # Mf-m
    if Sex == '1':
        temp_t = [4, 25, 69, 70, 74, 77, 78, 87, 92, 126, 132, 134, 140, 149, 179, 187, 203, 204, 217, 226, 231, 239,
                  261, 278, 282, 295, 297, 299]
        temp_f = [1, 19, 26, 28, 79, 80, 81, 89, 99, 112, 115, 116, 117, 120, 133, 144, 176, 198, 213, 214, 219, 221,
                  223, 229, 249, 254, 260, 262, 264, 280, 283, 300]
    # feminine masculinity
    # Mf-f
    else:
        temp_t = [4, 25, 70, 74, 77, 78, 87, 92, 126, 132, 133, 134, 140, 149, 187, 203, 204, 217, 226, 239, 261, 278,
                  282, 295, 299]
        temp_f = [1, 19, 26, 28, 69, 79, 80, 81, 89, 99, 112, 115, 116, 117, 120, 144, 176, 179, 198, 213, 214, 219,
                  221, 223, 229, 231, 249, 254, 260, 262, 264, 280, 283, 297, 300]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Mf'], Norm_SD['Mf'])

    return ori_score, pro_score


def scale_pa(ori_score=0, pro_score=0):
    """
    Clinical Scale-6 Paranoia Pa
    the score of Pa (paranoia) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [16, 24, 27, 35, 110, 121, 123, 127, 151, 157, 158, 202, 275, 284, 291, 293, 299, 305, 314, 317, 326, 338,
              341, 364, 365]
    temp_f = [93, 107, 109, 111, 117, 124, 268, 281, 294, 313, 316, 319, 327, 347, 348]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Pa'], Norm_SD['Pa'])

    return ori_score, pro_score


def scale_pt(ori_score=0, pro_score=0, pro_score_add_k=0):
    """
    Clinical Scale-7 Mental Asthenia Pt
    the score of Pt (psychasthenia) scale

    :param ori_score: original score
    :param pro_score: processing score
    :param pro_score_add_k: processing score added 1.0K
    :return: ori_score, pro_score, pro_score_add_k

    :rtype: int, int, int
    """
    temp_t = [10, 15, 22, 32, 41, 67, 76, 86, 94, 102, 106, 142, 159, 182, 189, 217, 238, 266, 301, 304, 321, 336, 337,
              340, 342, 343, 344, 346, 349, 351, 352, 356, 357, 358, 359, 360, 361, 362, 366]
    temp_f = [3, 8, 36, 122, 152, 164, 178, 329, 353]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    k, ignore = scale_k()
    pro_score += trans_t(ori_score, Norm_M['Pt'], Norm_SD['Pt'])
    pro_score_add_k += trans_t(ori_score + k, Norm_M['Pt+1.0K'], Norm_SD['Pt+1.0K'])

    return ori_score, pro_score, pro_score_add_k


def scale_sc(ori_score=0, pro_score=0, pro_score_add_k=0):
    """
    Clinical Scale-8 Schizophrenia Sc
    the score of Sc (schizophrenia) scale

    :param ori_score: original score
    :param pro_score: processing score
    :param pro_score_add_k: processing score added 1.0K
    :return: ori_score, pro_score, pro_score_add_k

    :rtype: int, int, int
    """
    temp_t = [15, 22, 40, 41, 47, 52, 76, 97, 104, 121, 156, 157, 159, 168, 179, 182, 194, 202, 210, 212, 238, 241, 251,
              259, 266, 273, 282, 291, 297, 301, 303, 307, 308, 311, 312, 315, 320, 323, 324, 325, 328, 331, 332, 333,
              334, 335, 339, 341, 345, 349, 350, 352, 354, 355, 356, 360, 363, 364, 366]
    temp_f = [17, 65, 103, 119, 177, 178, 187, 192, 196, 220, 276, 281, 302, 306, 309, 310, 318, 322, 330]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    k, ignore = scale_k()
    pro_score += trans_t(ori_score, Norm_M['Sc'], Norm_SD['Sc'])
    pro_score_add_k += trans_t(ori_score + k, Norm_M['Sc+1.0K'], Norm_SD['Sc+1.0K'])

    return ori_score, pro_score, pro_score_add_k


def scale_ma(ori_score=0, pro_score=0, pro_score_add_k=0):
    """
    Clinical Scale-9 Hypomania Ma
    the score of Ma (hypomania) scale

    :param ori_score: original score
    :param pro_score: processing score
    :param pro_score_add_k: processing score added 0.2K
    :return: ori_score, pro_score, pro_score_add_k

    :rtype: int, int, int
    """
    temp_t = [11, 13, 21, 22, 59, 64, 73, 97, 100, 109, 127, 134, 143, 156, 157, 167, 181, 194, 212, 222, 226, 228, 232,
              233, 238, 240, 250, 251, 263, 266, 268, 271, 277, 279, 298]
    temp_f = [101, 105, 111, 119, 120, 148, 166, 171, 180, 267, 289]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    k, ignore = scale_k()
    pro_score += trans_t(ori_score, Norm_M['Ma'], Norm_SD['Ma'])
    pro_score_add_k += trans_t(ori_score + round(0.2 * k), Norm_M['Ma+0.2K'], Norm_SD['Ma+0.2K'])

    return ori_score, pro_score, pro_score_add_k


def scale_si(ori_score=0, pro_score=0):
    """
    Clinical Scale-0 Social Introversion Si
    the score of Si (social introversion) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [32, 67, 82, 111, 117, 124, 138, 147, 171, 172, 180, 201, 236, 267, 278, 292, 304, 316, 321, 332, 336, 342,
              357, 369, 370, 373, 376, 378, 379, 385, 389, 393, 398, 399]
    temp_f = [25, 33, 57, 91, 99, 110, 126, 143, 193, 208, 229, 231, 254, 262, 281, 296, 309, 353, 359, 367, 371, 374,
              377, 380, 381, 382, 383, 384, 387, 388, 390, 391, 392, 395, 396, 397]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Si'], Norm_SD['Si'])

    return ori_score, pro_score


def scale_mas(ori_score=0, pro_score=0):
    """
    Additional Scale - Explicit Anxiety Mas
    the score of Mas (Manifest anxiety) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [13, 14, 23, 31, 32, 43, 67, 86, 125, 142, 158, 186, 191, 217, 238, 241, 263, 301, 317, 321, 322, 335, 337,
              340, 352, 361, 372, 398, 418, 424, 431, 439, 442, 499, 506, 530, 555]
    temp_f = [7, 18, 107, 163, 190, 230, 242, 264, 287, 367, 407, 520, 528]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Mas'], Norm_SD['Mas'])

    return ori_score, pro_score


def scale_dy(ori_score=0, pro_score=0):
    """
    Additional Scale - Dependency Dy
    the score of Dy (Dependency) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [19, 21, 24, 41, 63, 67, 70, 82, 86, 98, 100, 138, 141, 158, 165, 180, 189, 201, 212, 236, 239, 259, 267,
              304, 305, 321, 337, 338, 343, 357, 361, 362, 370, 372, 373, 393, 398, 399, 408, 440, 443, 461, 487, 488,
              489, 509, 521, 531, 554]
    temp_f = [9, 79, 107, 163, 170, 193, 264, 411]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Dy'], Norm_SD['Dy'])

    return ori_score, pro_score


def scale_do(ori_score=0, pro_score=0):
    """
    Additional Scale - Dominance Do
    the score of Do (Dominance) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [64, 229, 255, 270, 406, 432, 523]
    temp_f = [32, 61, 82, 86, 94, 186, 223, 224, 240, 249, 250, 267, 268, 304, 343, 356, 419, 483, 547, 558, 562]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Do'], Norm_SD['Do'])

    return ori_score, pro_score


def scale_re(ori_score=0, pro_score=0):
    """
    Additional Scale - Social Responsibility Re
    the score of Re (Social Responsibility) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [58, 111, 173, 221, 294, 412, 501, 552]
    temp_f = [6, 28, 30, 33, 56, 116, 118, 157, 175, 181, 223, 224, 260, 304, 388, 419, 434, 437, 468, 471, 472, 529,
              553, 558]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Re'], Norm_SD['Re'])

    return ori_score, pro_score


def scale_cn(ori_score=0, pro_score=0):
    """
    Additional Scale - Control Cn
    the score of Cn (Control) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [6, 20, 30, 56, 67, 105, 116, 134, 145, 162, 169, 181, 225, 236, 238, 285, 296, 319, 337, 376, 379, 381,
              418, 447, 460, 461, 529, 555]
    temp_f = [58, 80, 92, 96, 111, 167, 174, 220, 242, 249, 250, 291, 313, 360, 439, 444, 449, 483, 488, 489, 527, 548]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Cn'], Norm_SD['Cn'])

    return ori_score, pro_score


def calculate_score():
    """
    Test Score Calculation
    calculate the score

    :return: None
    """
    global ori_Point
    global pro_Point

    ori_Point = {

    }
    pro_Point = {

    }

    norm_select(Sex)

    ori_Point['Q*'], pro_Point['Q*'] = scale_q()
    ori_Point['L'], pro_Point['L'] = scale_l()
    ori_Point['F'], pro_Point['F'] = scale_f()
    ori_Point['K'], pro_Point['K'] = scale_k()
    ori_Point['Hs'], pro_Point['Hs'], pro_Point['Hs+0.5K'] = scale_hs()
    ori_Point['D'], pro_Point['D'] = scale_d()
    ori_Point['Hy'], pro_Point['Hy'] = scale_hy()
    ori_Point['Pd'], pro_Point['Pd'], pro_Point['Pd+0.4K'] = scale_pd()
    ori_Point['Mf'], pro_Point['Mf'] = scale_mf()
    ori_Point['Pa'], pro_Point['Pa'] = scale_pa()
    ori_Point['Pt'], pro_Point['Pt'], pro_Point['Pt+1.0K'] = scale_pt()
    ori_Point['Sc'], pro_Point['Sc'], pro_Point['Sc+1.0K'] = scale_sc()
    ori_Point['Ma'], pro_Point['Ma'], pro_Point['Ma+0.2K'] = scale_ma()
    ori_Point['Si'], pro_Point['Si'] = scale_si()
    ori_Point['Mas'], pro_Point['Mas'] = scale_mas()
    ori_Point['Dy'], pro_Point['Dy'] = scale_dy()
    ori_Point['Do'], pro_Point['Do'] = scale_do()
    ori_Point['Re'], pro_Point['Re'] = scale_re()
    ori_Point['Cn'], pro_Point['Cn'] = scale_cn()


def analyze_score():
    """
    Analyzing Test Scores
    analyze the score of test

    Use the two-point coding method and the analysis chart to present the personality characteristics of the subjects
    Use 2 point codes and personality profile to show the personality traits of the subjects

    :return:None
    """
    global two_point

    val_scale = ['L', 'F', 'K']
    cli_scale = ['Hs\n1', 'D\n2', 'Hy\n3', 'Pd\n4', 'Mf\n5', 'Pa\n6', 'Pt\n7', 'Sc\n8', 'Ma\n9', 'Si\n0']
    ext_scale = ['Mas', 'Dy', 'Do', 'Re', 'Cn']

    val_list = [
        pro_Point['L'],
        pro_Point['F'],
        pro_Point['K']
    ]
    cli_list = [
        pro_Point['Hs+0.5K'],
        pro_Point['D'],
        pro_Point['Hy'],
        pro_Point['Pd+0.4K'],
        pro_Point['Mf'],
        pro_Point['Pa'],
        pro_Point['Pt+1.0K'],
        pro_Point['Sc+1.0K'],
        pro_Point['Ma+0.2K'],
        pro_Point['Si']
    ]
    ext_list = [
        pro_Point['Mas'],
        pro_Point['Dy'],
        pro_Point['Do'],
        pro_Point['Re'],
        pro_Point['Cn']
    ]

    cli_max1 = max(cli_list)
    cli_max1_index = cli_list.index(cli_max1)
    if cli_max1_index != 9:
        first = cli_max1_index + 1
    else:
        first = 0
    cli_list[cli_max1_index] = 0
    cli_max2 = max(cli_list)
    cli_max2_index = cli_list.index(cli_max2)
    if cli_max2_index != 9:
        second = cli_max2_index + 1
    else:
        second = 0
    cli_list[cli_max1_index] = cli_max1
    two_point = '%s%s' % (str(first), str(second))

    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False
    fig = plt.figure(figsize=(10, 6), dpi=100, linewidth=1)
    ax = fig.add_subplot(111)
    ax.plot(range(len(val_list)), val_list, 'b*-')
    ax.plot(range(len(val_list), len(val_list) + len(cli_list)), cli_list, 'b*-')
    ax.plot(range(len(val_list) + len(cli_list), len(val_list) + len(cli_list) + len(ext_list)), ext_list, 'b*-')
    # plt.setp(ax.xaxis.get_majorticklabels(), rotation=-45)
    ax.set_xticks(range(len(val_list + cli_list + ext_list)))
    ax.set_xticklabels(val_scale + cli_scale + ext_scale)
    ax.set_yticks([0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 110, 120])
    ax.set_xlim(-0.5, len(val_scale + cli_scale + ext_scale) - 0.5)
    ax.set_ylim(0, 120)
    plt.axvline(2.5, ls="-", color="black")
    plt.axvline(12.5, ls="-", color="black")
    plt.axvline(7, ls="--", color="red")
    plt.axhline(50, ls="-", color="black")
    plt.axhline(60, ls="--", color="black")
    plt.axhline(70, ls="-", color="black")
    ax.plot(cli_max1_index + len(val_list), cli_max1, 'rp')
    ax.plot(cli_max2_index + len(val_list), cli_max2, 'rp')
    plt.annotate(r'$max1$', xy=(cli_max1_index + len(val_list)+0.1, cli_max1+2), color='red', fontsize=8)
    plt.annotate(r'$max2$', xy=(cli_max2_index + len(val_list)+0.1, cli_max2+2), color='red', fontsize=8)


def data_export():
    """
    export data

    export test data

    :return: None

    Note: Generate a '.xlsx' file to save the test information
    """
    print('Please enter the subject's name')
    name = input('> ')
    wb = Workbook()
    data_filename = time.strftime("%Y%m%d_%H%M_", time.localtime()) + name + '_MMPI-test'

    font1 = Font(name='Times New Roman', size=12)
    font2 = Font(name='Times New Roman', size=12)
    font3 = Font(name='Times New Roman', size=12, bold=True)
    font4 = Font(name='Times New Roman', size=12)
    alig1 = Alignment(horizontal='center', vertical='center')
    alig2 = Alignment(horizontal='general', vertical='center')

    # Table 1. Raw data of record test
    sheet1 = wb.active
    sheet1.title = 'quiz raw data'
    sheet1['A1'] = 'Name'
    sheet1['A1'].font = font1
    sheet1['A1'].alignment = alig1
    sheet1['C1'] = 'gender'
    sheet1['C1'].font = font1
    sheet1['C1'].alignment = alig1
    sheet1['E1'] = 'age'
    sheet1['E1'].font = font1
    sheet1['E1'].alignment = alig1
    sheet1.merge_cells('A2:B2')
    sheet1['A2'] = 'Topic'
    sheet1['A2'].font = font1
    sheet1['A2'].alignment = alig1
    sheet1.merge_cells('C2:D2')
    sheet1['C2'] = 'answer'
    sheet1['C2'].font = font1
    sheet1['C2'].alignment = alig1
    sheet1['B1'] = name
    sheet1['B1'].font = font2
    sheet1['B1'].alignment = alig1
    if Sex == '1':
        sex_name = 'male'
    else:
        sex_name = 'female'
    sheet1['D1'] = sex_name
    sheet1['D1'].font = font2
    sheet1['D1'].alignment = alig1
    sheet1['F1'] = Age
    sheet1['F1'].font = font4
    sheet1['F1'].alignment = alig1
    for i in range(len(Que)+1):
        if i == 73:
            if Sex == '1':
                temp_que = Que[i+1][Que[i+1].find('m')+1: Que[i+1].find('f')]
            else:
                temp_que = Que[i+1][Que[i+1].find('f') + 1:]
        elif i == len(Que):
            temp_que = 'I promise to complete this quiz seriously and honestly under the guidance of professionals'
        else:
            temp_que = Que[i+1]

        sheet1['A%d' % (i + 3)].value = str(i + 1) + '.'
        sheet1['A%d' % (i + 3)].font = font2
        sheet1['A%d' % (i + 3)].alignment = alig1
        sheet1['B%d' % (i + 3)].value = temp_que
        sheet1['B%d' % (i + 3)].font = font2
        sheet1['B%d' % (i + 3)].alignment = alig2
        if Ans[i+1] == '1':
            temp_ans = '是'
            sheet1['C%d' % (i + 3)].value = temp_ans
            sheet1['C%d' % (i + 3)].font = font2
            sheet1['C%d' % (i + 3)].alignment = alig1
        else:
            temp_ans = '否'
            sheet1['D%d' % (i + 3)].value = temp_ans
            sheet1['D%d' % (i + 3)].font = font2
            sheet1['D%d' % (i + 3)].alignment = alig1

    # Table 2, Recording Test Scores
    sheet2 = wb.create_sheet(title='test score')
    sheet2['A1'] = 'Name'
    sheet2['A1'].font = font1
    sheet2['A1'].alignment = alig1
    sheet2['C1'] = 'gender'
    sheet2['C1'].font = font1
    sheet2['C1'].alignment = alig1
    sheet2['E1'] = 'age'
    sheet2['E1'].font = font1
    sheet2['E1'].alignment = alig1
    sheet2['B1'].value = name
    sheet2['B1'].font = font2
    sheet2['B1'].alignment = alig1
    sheet2['D1'].value = sex_name
    sheet2['D1'].font = font2
    sheet2['D1'].alignment = alig1
    sheet2['F1'].value = Age
    sheet2['F1'].font = font4
    sheet2['F1'].alignment = alig1
    sheet2['A2'] = 'Scale category'
    sheet2['A2'].font = font1
    sheet2['A2'].alignment = alig1
    sheet2['B2'] = 'raw score'
    sheet2['B2'].font = font1
    sheet2['B2'].alignment = alig1
    sheet2['C2'] = 'Standard score (without K) '
    sheet2['C2'].font = font1
    sheet2['C2'].alignment = alig1
    sheet2['D2'] = 'Standard score (plus K) '
    sheet2['D2'].font = font1
    sheet2['D2'].alignment = alig1
    sheet2['A3'] = 'The Q* scale only records the number of contradictory questions'
    sheet2['A3'].font = font2
    sheet2['A3'].alignment = alig2
    scale_list = ['Q*', 'L', 'F', 'K', 'Hs', 'D', 'Hy', 'Pd', 'Mf', 'Pa', 'Pt', 'Sc', 'Ma', 'Si',
                  'Mas', 'Dy', 'Do', 'Re', 'Cn']
    for i in range(len(scale_list)):
        sheet2['A%d' % (i+4)].value = scale_list[i]
        sheet2['A%d' % (i+4)].font = font3
        sheet2['A%d' % (i+4)].alignment = alig1
        sheet2['B%d' % (i+4)].value = ori_Point[scale_list[i]]
        sheet2['B%d' % (i+4)].font = font4
        sheet2['B%d' % (i+4)].alignment = alig1
        sheet2['C%d' % (i+4)].value = pro_Point[scale_list[i]]
        sheet2['C%d' % (i+4)].font = font4
        sheet2['C%d' % (i+4)].alignment = alig1
        if scale_list[i] == 'Hs':
            sheet2['D%d' % (i + 4)].value = pro_Point['Hs+0.5K']
            sheet2['D%d' % (i + 4)].font = font4
            sheet2['D%d' % (i + 4)].alignment = alig1
        elif scale_list[i] == 'Pd':
            sheet2['D%d' % (i + 4)].value = pro_Point['Pd+0.4K']
            sheet2['D%d' % (i + 4)].font = font4
            sheet2['D%d' % (i + 4)].alignment = alig1
        elif scale_list[i] == 'Pt':
            sheet2['D%d' % (i + 4)].value = pro_Point['Pt+1.0K']
            sheet2['D%d' % (i + 4)].font = font4
            sheet2['D%d' % (i + 4)].alignment = alig1
        elif scale_list[i] == 'Sc':
            sheet2['D%d' % (i + 4)].value = pro_Point['Sc+1.0K']
            sheet2['D%d' % (i + 4)].font = font4
            sheet2['D%d' % (i + 4)].alignment = alig1
        elif scale_list[i] == 'Ma':
            sheet2['D%d' % (i + 4)].value = pro_Point['Ma+0.2K']
            sheet2['D%d' % (i + 4)].font = font4
            sheet2['D%d' % (i + 4)].alignment = alig1
        else:
            pass
    sheet2['E2'] = 'two point encoding'
    sheet2['E2'].font = font1
    sheet2['E2'].alignment = alig1
    sheet2['F2'].value = two_point
    sheet2['F2'].font = font4
    sheet2['F2'].alignment = alig1

    wb.save(filename=data_filename + '.xlsx')
    plt.title('%s MAnatomical diagram Add K points to correct T points (Chinese norm)' % name)
    plt.savefig(data_filename)
